

from openpyxl import load_workbook

class ExcelHandler():

    def __init__(self, file_name, sheet_name):

        self.file_name = file_name
        self.sheet_name = sheet_name

    def read_data(self):
        '''读取excel数据'''
        cases = []  # 存储所有用例
        titles = []  # 存储用例表头
        workbook = load_workbook(self.file_name)  # 获取Excel文件
        worksheet = workbook[self.sheet_name]  #获取sheet
        # 遍历所有行
        # 如果是表头，把表头保存到titles列表中，否则把数据添加到cases列表中
        for i, row in enumerate(worksheet.rows):
            if i == 0:
                for cell in row:
                    titles.append(cell.value)
            else:
                cases.append(dict(zip(titles, [cell.value for cell in row])))
        return cases

    def write_data(self, row, column, value):
        """数据写入指定单元格
        args:
            row: 行号
            column: 列号
            value:  要写入的值
        returns:
            None
        """
        workbook = load_workbook(self.file_name)  # 获取一个WorkBook对象
        worksheet = workbook[self.sheet_name]  # 获取一个WorkSheet对象
        worksheet.cell(row, column, value)
        workbook.save(self.file_name)
        return None
